#Busca conferencias sobre LSTI que ya predetermine y busca informacion del clima en el cual sera dicha conferencia asi como
#su informacion de redes sociales y se guarda todo en un archivo de excel
import os, requests, re, xlrd, json
from googlesearch import search
from bs4 import BeautifulSoup as bs
from openpyxl import Workbook

#Pagina de prueba http://sitios.itesm.mx/cite/info.htm

def buscarPagina():
	idLibro = 1
	listaArchivos = os.listdir('.')
	
	
	libro = Workbook()
	hoja = libro.active
	hoja.cell(row = 1, column = 1).value='EVENTOS'
	hoja.cell(row = 2, column = 1).value='Informacion del evento'
	hoja.cell(row = 2, column = 4).value='Contacto'
	hoja.cell(row = 2, column = 9).value='Clima'

	c = 4 #Empezamos aqui (info contacto) ya que no conseguimos los datos del evento
	f = 3

	while True:
		#Empezamos aqui (info contacto) ya que no conseguimos los datos del evento
		c = 4
		selec = 0
		valor = 0
		contador = 0
		lista = []
		listaArchivos = os.listdir('.')
		query = input('\nBusqueda:\t\t\t\t\tSalir (s)\n> ')

		if query == 's':
			
			while True:
				if 'libroEventos' + str(idLibro) + '.xlsx' in listaArchivos:
					idLibro +=1
				else:
					break
			libro.save('libroEventos' + str(idLibro) + '.xlsx')
			print('\nguardando informacion en el libroEventos' + str(idLibro) + '.xlsx ...\n\n[OK]\n')
			break

		print('\nBuscando...')
		
		for enlace in search(query, tld="com", num=15, stop=15, pause=5):
			print('url: ' + enlace + '\n')	
			if 'Direcciones.web.txt' in listaArchivos:
				ft = open('Direcciones.web.txt', 'a', encoding="utf-8")
				ft.write('\n' + enlace) 
			else:
				ft = open('Direcciones.web.txt', 'w', encoding="utf-8")
				ft.write(enlace) 
			ft.close()
			pagina =  requests.get(enlace) 
			if pagina.status_code != 200:
				print('Pagina no encontrada!')
				break
			print('descargando pagina...\n')

			
			for archivo in listaArchivos:
				if os.path.isdir(os.path.join('.', archivo)):
					contador +=1
			contador +=1
			#print(listaArchivos)
			#print(contador)
			#input()
			print('creando directorio PAGINA' + str(contador) + '...\n')
			os.makedirs('PAGINA' + str(contador))
			soup = bs(pagina.content, 'html.parser')        
			ruta = os.getcwd() + '/PAGINA' + str(contador) + '/'
			print('creando html en el nuevo directorio...\n')
			fh = open(ruta + 'pagina.html', 'w', encoding="utf-8")
			contenido = soup.prettify()
			fh.write(contenido)
			fh.close()
			fh = open(ruta + 'pagina.html', 'r', encoding="utf-8")
			correo = re.compile(r'(\")([\w:.\-\_]+@[a-z.]+)(\")')
			facebook = re.compile(r'(\")([a-z:/.]+facebook[a-zA-Z0-9:/.?=\-]+)(\")')
			twitter = re.compile(r'(\")([a-z:/]+twitter[a-z./?=_]+)(\")')
			instagram = re.compile(r'(\")([a-z:/.]+instagram[a-z/.]+)')
			for line in fh:	
				mo1 = correo.search(line)
				mo2 = facebook.search(line)
				mo3 = twitter.search(line)
				mo4 = instagram.search(line)
				if not mo1 == None:
					lista.append(mo1.group(2))
				if not mo2 == None:
					lista.append(mo2.group(2))
				if not mo3 == None:
					lista.append(mo3.group(2))
				if not mo4 == None:
					lista.append(mo4.group(2))

			page1 = requests.get ("https://api.openweathermap.org/data/2.5/onecall?lat=25.725536&lon=-100.315157&exclude=minutely,hourly,daily&units=metric&appid=41b9652ca4fc11607d3571bc81bfb42c")    
			weatherData1 = str(json.loads(page1.content))
			lista.append(weatherData1)

			#print(lista)
			
			
			for elemento in lista:
				hoja.cell(row = f, column = c).value = elemento
				c +=1
			f +=1
			break
			
			
			#page2 = requests.get ("https://api.openweathermap.org/data/2.5/onecall?lat=-33.4372&lon=-70.6506&exclude=minutely,hourly,daily&units=metric&appid=41b9652ca4fc11607d3571bc81bfb42c")    
			#weatherData2 =  str(json.loads(page2.content))
			#lista.append(weatherData2)
			#page3 = requests.get ("https://api.openweathermap.org/data/2.5/onecall?lat=46.2276382&lon=2.2137489&exclude=minutely,hourly,daily&units=metric&appid=41b9652ca4fc11607d3571bc81bfb42c")    
			#weatherData3 =  str(json.loads(page3.content))
			#lista.append(weatherData3)
			#page4 = requests.get ("https://api.openweathermap.org/data/2.5/onecall?lat=19.4284706&lon=-99.1276627&exclude=minutely,hourly,daily&units=metric&appid=41b9652ca4fc11607d3571bc81bfb42c")    
			#weatherData4 =  str(json.loads(page4.content))
			#lista.append(weatherData4)
			#page5 = requests.get ("https://api.openweathermap.org/data/2.5/onecall?lat=40.4167&lon=-3.70325&exclude=minutely,hourly,daily&units=metric&appid=41b9652ca4fc11607d3571bc81bfb42c")    
			#weatherData5 =  str(json.loads(page5.content))
			#lista.append(weatherData5)
			#page6 = requests.get ("https://api.openweathermap.org/data/2.5/onecall?lat=4.6097100&lon=-74.0817500&exclude=minutely,hourly,daily&units=metric&appid=41b9652ca4fc11607d3571bc81bfb42c")    
			#weatherData6 =  str(json.loads(page6.content))
			#lista.append(weatherData6)
			#page7 = requests.get ("https://api.openweathermap.org/data/2.5/onecall?lat=21.1236&lon=-101.68&exclude=minutely,hourly,daily&units=metric&appid=41b9652ca4fc11607d3571bc81bfb42c")    
			#weatherData7 =  str(json.loads(page7.content))
			#lista.append(weatherData7)
			
			
			
			
                
def abrirArchivoTexto():
	idLibro = 1	
	libro = Workbook()
	hoja = libro.active
	hoja.cell(row = 1, column = 1).value='EVENTOS'
	hoja.cell(row = 2, column = 1).value='Informacion del evento'
	hoja.cell(row = 2, column = 4).value='Contacto'
	hoja.cell(row = 2, column = 9).value='Clima'

	#Empezamos aqui (info contacto) ya que no conseguimos los datos del evento
	f = 3

	dirpath = '.'
	listaArchivos = os.listdir(dirpath)
	lista = []
	idEnlace = 1

	print('\nArchivos en directorio:\n', listaArchivos, '\n')
	nameFile = str(input('Ingrese el nombre del archivo que desea abrir\n> '))
	if not nameFile in listaArchivos:
		print('Archivo no encontrado!')
	else:
		ft = open(nameFile, 'r', encoding="utf-8") 
		contador = 0
		for enlace in ft:
			c = 4
			print('\ndescargando ' + str(idEnlace)+ ' ' + enlace + '\n')
			pagina = requests.get(enlace) 
			
			if pagina.status_code != 200:
				print('Pagina no encontrada!') 
				print(pagina.status_code)

			else:
				soup = bs(pagina.content, 'html.parser') 
				for archivo in listaArchivos:
					if os.path.isdir(os.path.join('.', archivo)):
						contador +=1
				contador +=1
				print('creando directorio PAGINA' + str(contador) + '...\n')
				os.makedirs('PAGINA' + str(contador)) 
		
				ruta = os.getcwd() + '/PAGINA' + str(contador) + '/'

				print('creando html...')
				fh = open(ruta + 'pagina.html', 'w', encoding="utf-8")
				contenido = soup.prettify()
				fh.write(contenido)
				fh.close()

				fh = open(ruta + 'pagina.html', 'r', encoding="utf-8")
		
				correo = re.compile(r'(\")([\w:.\-\_]+@[a-z.]+)(\")')
				facebook = re.compile(r'(\")([a-z:/.]+facebook[a-zA-Z0-9:/.?=\-]+)(\")')
				twitter = re.compile(r'(\")([a-z:/]+twitter[a-z./?=_]+)(\")')
				instagram = re.compile(r'(\")([a-z:/.]+instagram[a-z/.]+)')

				for line in fh:

					mo1 = correo.search(line)
					mo2 = facebook.search(line)
					mo3 = twitter.search(line)
					mo4 = instagram.search(line)

					if not mo1 == None:
						lista.append(mo1.group(2))
					if not mo2 == None:
						lista.append(mo2.group(2))
					if not mo3 == None:
						lista.append(mo3.group(2))
					if not mo4 == None:
						lista.append(mo4.group(2))

				page1 = requests.get ("https://api.openweathermap.org/data/2.5/onecall?lat=25.725536&lon=-100.315157&exclude=minutely,hourly,daily&units=metric&appid=41b9652ca4fc11607d3571bc81bfb42c")    
				weatherData1 = str(json.loads(page1.content))
				lista.append(weatherData1)
				
				
				for elemento in lista:
					hoja.cell(row = f, column = c).value = elemento
					c +=1
				f +=1
			idEnlace +=1
				
			
	while True:
		if 'libroEventos' + str(idLibro) + '.xlsx' in listaArchivos:
			idLibro +=1
		else:
			break
	libro.save('libroEventos' + str(idLibro) + '.xlsx')
	print('\nguardando informacion en el libroEventos' + str(idLibro) + '.xlsx ...\n\n[OK]\n')
			


while True:

	print('Seleccione una opcion:\n Buscar en la web (1)\n Abrir archivo de texto (2)\n Salir (3)')
	op = str(input('> '))

	if op == '1':
		buscarPagina()
	
	if op == '2':
		abrirArchivoTexto()

	if op == '3':
		print('\nsaliendo...')
		break


	if op != '1' and op != '2' and op != '3':
		print('Error, el valor ingresado no es valido')

'''
print("\nHora en nuestra zona horaria.\n")

horamty = tzlocal.get_localzone()
print(horamty)

hora = datetime.fromtimestamp(weatherData1["current"]["dt"],horamty)
print("dt: ",hora.strftime("%d/%m/%Y %H:%M"))
hora = datetime.fromtimestamp(weatherData2["current"]["dt"],horamty)
print("dt: ",hora.strftime("%d/%m/%Y %H:%M")) 
hora = datetime.fromtimestamp(weatherData3["current"]["dt"],horamty)
print("dt: ",hora.strftime("%d/%m/%Y %H:%M")) 
hora = datetime.fromtimestamp(weatherData4["current"]["dt"],horamty)
print("dt: ",hora.strftime("%d/%m/%Y %H:%M")) 
hora = datetime.fromtimestamp(weatherData5["current"]["dt"],horamty)
print("dt: ",hora.strftime("%d/%m/%Y %H:%M")) 
hora = datetime.fromtimestamp(weatherData6["current"]["dt"],horamty)
print("dt: ",hora.strftime("%d/%m/%Y %H:%M")) 
hora = datetime.fromtimestamp(weatherData7["current"]["dt"],horamty)
print("dt: ",hora.strftime("%d/%m/%Y %H:%M")) 
hora = datetime.fromtimestamp(weatherData8["current"]["dt"],horamty)
print("dt: ",hora.strftime("%d/%m/%Y %H:%M")) 


sunrise = datetime.fromtimestamp(weatherData1["current"]["sunrise"],horamty)
print("sunrise(amanecer): ",sunrise.strftime("%d/%m/%Y %H:%M:%S"))

sunrise = datetime.fromtimestamp(weatherData2["current"]["sunrise"],horamty)
print("sunrise(amanecer): ",sunrise.strftime("%d/%m/%Y %H:%M:%S"))
sunrise = datetime.fromtimestamp(weatherData3["current"]["sunrise"],horamty)
print("sunrise(amanecer): ",sunrise.strftime("%d/%m/%Y %H:%M:%S"))
sunrise = datetime.fromtimestamp(weatherData4["current"]["sunrise"],horamty)
print("sunrise(amanecer): ",sunrise.strftime("%d/%m/%Y %H:%M:%S"))
sunrise = datetime.fromtimestamp(weatherData5["current"]["sunrise"],horamty)
print("sunrise(amanecer): ",sunrise.strftime("%d/%m/%Y %H:%M:%S"))
sunrise = datetime.fromtimestamp(weatherData6["current"]["sunrise"],horamty)
print("sunrise(amanecer): ",sunrise.strftime("%d/%m/%Y %H:%M:%S"))
sunrise = datetime.fromtimestamp(weatherData7["current"]["sunrise"],horamty)
print("sunrise(amanecer): ",sunrise.strftime("%d/%m/%Y %H:%M:%S"))
sunrise = datetime.fromtimestamp(weatherData8["current"]["sunrise"],horamty)
print("sunrise(amanecer): ",sunrise.strftime("%d/%m/%Y %H:%M:%S"))

sunset = datetime.fromtimestamp(weatherData1["current"]["sunset"],horamty)
print("sunset(atarceder): ",sunset.strftime("%d/%m/%Y %H:%M:%S"))
sunset = datetime.fromtimestamp(weatherData2["current"]["sunset"],horamty)
print("sunset(atarceder): ",sunset.strftime("%d/%m/%Y %H:%M:%S"))
sunset = datetime.fromtimestamp(weatherData3["current"]["sunset"],horamty)
print("sunset(atarceder): ",sunset.strftime("%d/%m/%Y %H:%M:%S"))
sunset = datetime.fromtimestamp(weatherData4["current"]["sunset"],horamty)
print("sunset(atarceder): ",sunset.strftime("%d/%m/%Y %H:%M:%S"))
sunset = datetime.fromtimestamp(weatherData5["current"]["sunset"],horamty)
print("sunset(atarceder): ",sunset.strftime("%d/%m/%Y %H:%M:%S"))
sunset = datetime.fromtimestamp(weatherData6["current"]["sunset"],horamty)
print("sunset(atarceder): ",sunset.strftime("%d/%m/%Y %H:%M:%S"))
sunset = datetime.fromtimestamp(weatherData7["current"]["sunset"],horamty)
print("sunset(atarceder): ",sunset.strftime("%d/%m/%Y %H:%M:%S"))
sunset = datetime.fromtimestamp(weatherData8["current"]["sunset"],horamty)
print("sunset(atarceder): ",sunset.strftime("%d/%m/%Y %H:%M:%S"))
'''











